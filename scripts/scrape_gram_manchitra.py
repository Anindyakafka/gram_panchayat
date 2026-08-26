#!/usr/bin/env python3
"""Extract selected Gram Manchitra GP profiles and election details."""

from __future__ import annotations

import argparse
import csv
import json
import re
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROFILE_URL = "https://grammanchitra.gov.in/gm4MVC/Home/GetPanchAreaDetails"
VILLAGE_URL = (
    "https://grammanchitragis.nic.in/grammanchitra/rest/services/"
    "panchayat/panchayat_admin/MapServer/7/query"
)
IDENTIFIERS = ["State", "District", "Block", "GP name", "Our GP ID"]
ELECTION_FIELDS = [
    ("wardName", "Ward name"),
    ("name", "Member/elected representative name"),
    ("wardType", "Ward type"),
    ("gender", "Gender"),
    ("caste", "Caste/category"),
    ("qualification", "Educational qualification"),
    ("age", "Age"),
    ("electionTermStart", "Election term start date"),
    ("electionTermEnddate", "Election term end date"),
    ("electionType", "Election type"),
]
BASIC_AUDIT_FIELDS = {
    "stateId", "distId", "subDistrict", "village", "localBodyCode",
    "nameOfSarpanch", "spgenderId", "nameOfSecretary", "psgenderId",
}


def fetch_json(request: urllib.request.Request, attempts: int = 4) -> dict:
    last_error = None
    for attempt in range(attempts):
        try:
            with urllib.request.urlopen(request, timeout=90) as response:
                return json.loads(response.read().decode("utf-8"))
        except Exception as error:
            last_error = error
            if attempt + 1 < attempts:
                time.sleep(2 ** attempt)
    raise RuntimeError(f"Request failed after {attempts} attempts: {request.full_url}") from last_error


def post_profile(gp_code: int, index: int) -> dict:
    payload = urllib.parse.urlencode({"index": index, "gpCode": gp_code}).encode()
    request = urllib.request.Request(PROFILE_URL, data=payload, method="POST")
    request.add_header("User-Agent", "Mozilla/5.0 Gram-Manchitra-research-extract/1.0")
    return fetch_json(request)


def get_villages(gp_code: int) -> dict:
    params = {
        "f": "json",
        "where": f"GP_CODE = {gp_code}",
        "outFields": "vilname,vilcode11,vilcode_n,GP_CODE,GP_NAME,Block_Name,dtname,stname",
        "returnGeometry": "false",
        "orderByFields": "vilname",
    }
    request = urllib.request.Request(VILLAGE_URL + "?" + urllib.parse.urlencode(params))
    request.add_header("User-Agent", "Mozilla/5.0 Gram-Manchitra-research-extract/1.0")
    return fetch_json(request)


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def read_gp_codes(path: Path) -> dict[str, set[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["gp_code"]
    rows = sheet.iter_rows(values_only=True)
    header = [norm(value) for value in next(rows)]
    code_col = next(i for i, value in enumerate(header) if value == "code")
    name_col = header.index("gp_name")
    matches: dict[str, set[str]] = {}
    for row in rows:
        if row[name_col] is None or row[code_col] is None:
            continue
        matches.setdefault(norm(row[name_col]), set()).add(str(row[code_col]).strip())
    return matches


def read_targets(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def hierarchy_from_villages(payload: dict) -> dict[str, str]:
    features = payload.get("features") or []
    if not features:
        raise ValueError("Village query returned no features")
    first = features[0]["attributes"]
    expected = {
        "State": first.get("stname"),
        "District": first.get("dtname"),
        "Block": first.get("Block_Name"),
        "GP name": first.get("GP_NAME"),
    }
    for feature in features[1:]:
        current = feature["attributes"]
        comparison = [current.get(k) for k in ("stname", "dtname", "Block_Name", "GP_NAME")]
        if [norm(v) for v in comparison] != [norm(v) for v in expected.values()]:
            raise ValueError("Village features disagree on the administrative hierarchy")
    return expected


def blank_if_none(value: object) -> object:
    return "" if value is None else value


def excel_date(value: object) -> object:
    if not value:
        return ""
    try:
        return date.fromisoformat(str(value))
    except ValueError:
        return value


def style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(1, sheet.max_row + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(map(len, values)) + 2, 45)


def save_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def run(source: Path, targets_path: Path, output: Path, raw_dir: Path) -> None:
    codes = read_gp_codes(source)
    targets = read_targets(targets_path)
    gp_records = []
    ward_records = []
    extra_election_keys: list[str] = []
    known_election_keys = {key for key, _ in ELECTION_FIELDS}

    for target in targets:
        gp_name = target["gp_name"]
        gp_code_matches = codes.get(norm(gp_name))
        if gp_code_matches is None:
            raise KeyError(f"No gp_code match for {gp_name!r}")
        if len(gp_code_matches) != 1:
            raise ValueError(f"Ambiguous gp_code matches for {gp_name!r}: {gp_code_matches}")
        gp_code_text = next(iter(gp_code_matches))
        gp_code = int(float(gp_code_text))
        basic = post_profile(gp_code, 0)
        glance = post_profile(gp_code, 5)
        villages_payload = get_villages(gp_code)
        hierarchy = hierarchy_from_villages(villages_payload)

        checks = {
            "State": target["state"],
            "District": target["district"],
            "Block": target["block"],
            "GP name": gp_name,
        }
        for field, expected in checks.items():
            if norm(hierarchy[field]) != norm(expected):
                raise ValueError(f"Hierarchy mismatch for {gp_name}: {field}={hierarchy[field]!r}")
        if int(glance["localBodyCode"]) != gp_code or norm(glance["localBodyName"]) != norm(gp_name):
            raise ValueError(f"Profile identity mismatch for {gp_name}")

        raw_basic = {key: value for key, value in basic.items() if key in BASIC_AUDIT_FIELDS}
        save_json(raw_dir / f"{gp_code}_{norm(gp_name).replace(' ', '_')}_basic.json", raw_basic)
        save_json(raw_dir / f"{gp_code}_{norm(gp_name).replace(' ', '_')}_glance.json", glance)
        save_json(raw_dir / f"{gp_code}_{norm(gp_name).replace(' ', '_')}_villages.json", villages_payload)

        villages = [item["attributes"].get("vilname") for item in villages_payload["features"]]
        notes = []
        if basic.get("spgenderId") is not None:
            notes.append(f"Sarpanch gender blank: site response only supplied unlabeled code {basic['spgenderId']}")
        else:
            notes.append("Sarpanch gender blank: not available")
        if basic.get("psgenderId") is not None:
            notes.append(f"Secretary gender blank: site response only supplied unlabeled code {basic['psgenderId']}")
        else:
            notes.append("Secretary gender blank: not available")
        gp_records.append({
            **checks,
            "Our GP ID": gp_code,
            "villages": villages,
            "Sarpanch name": blank_if_none(basic.get("nameOfSarpanch")),
            "Sarpanch gender": "",
            "Secretary name": blank_if_none(basic.get("nameOfSecretary")),
            "Secretary gender": "",
            "Notes": "; ".join(notes),
        })

        for election in glance.get("electionDetails") or []:
            record = {**checks, "Our GP ID": gp_code}
            missing = []
            for key, label in ELECTION_FIELDS:
                value = election.get(key)
                # The live table treats zero age as unavailable and renders N/A.
                unavailable = value is None or (key == "age" and value == 0)
                record[label] = "" if unavailable else (
                    excel_date(value) if "date" in label.casefold() else value
                )
                if unavailable:
                    missing.append(label)
            for key, value in election.items():
                if key not in known_election_keys:
                    if key not in extra_election_keys:
                        extra_election_keys.append(key)
                    record[key] = blank_if_none(value)
            record["Notes"] = "Blank on Gram Manchitra: " + ", ".join(missing) if missing else ""
            ward_records.append(record)

    max_villages = max((len(record["villages"]) for record in gp_records), default=0)
    gp_headers = IDENTIFIERS + [f"Village {i}" for i in range(1, max_villages + 1)] + [
        "Sarpanch name", "Sarpanch gender", "Secretary name", "Secretary gender", "Notes"
    ]
    ward_headers = (
        IDENTIFIERS
        + [label for _, label in ELECTION_FIELDS]
        + extra_election_keys
        + ["Notes"]
    )

    workbook = Workbook()
    gp_sheet = workbook.active
    gp_sheet.title = "GP Information"
    gp_sheet.append(gp_headers)
    for record in gp_records:
        row = [record[field] for field in IDENTIFIERS]
        row.extend(record["villages"] + [""] * (max_villages - len(record["villages"])))
        row.extend(record[field] for field in gp_headers[-5:])
        gp_sheet.append(row)

    ward_sheet = workbook.create_sheet("Ward Election Details")
    ward_sheet.append(ward_headers)
    for record in ward_records:
        ward_sheet.append([record.get(field, "") for field in ward_headers])
    for row in ward_sheet.iter_rows(min_row=2):
        for header, cell in zip(ward_headers, row):
            if "date" in header.casefold() and isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"

    style_sheet(gp_sheet)
    style_sheet(ward_sheet)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(f"Wrote {output} with {len(gp_records)} GPs and {len(ward_records)} election rows")


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=root / "gram_panchayat_reservation.xlsx")
    parser.add_argument("--targets", type=Path, default=root / "config" / "targets.csv")
    parser.add_argument("--output", type=Path, default=root / "output" / "gram_manchitra_bhandarda_dhamnai.xlsx")
    parser.add_argument("--raw-dir", type=Path, default=root / "data" / "raw")
    args = parser.parse_args()
    run(args.source, args.targets, args.output, args.raw_dir)


if __name__ == "__main__":
    main()
